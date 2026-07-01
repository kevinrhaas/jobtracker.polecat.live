import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

data=json.load(open('jobtracker_data.json'))
fields=data['fields']; records=data['records']; views=data['views']
fnames=[f['name'] for f in fields]

HEAD=Font(name='Arial',bold=True,color='FFFFFF',size=10)
BODY=Font(name='Arial',size=10)
FILL=PatternFill('solid',fgColor='2F5496')
ALT=PatternFill('solid',fgColor='EEF2F9')
WRAP=Alignment(vertical='top',wrap_text=True)
TOP=Alignment(vertical='top')
thin=Side(style='thin',color='D9D9D9')
BORD=Border(bottom=thin)

def render(name,val):
    if val is None: return ''
    if isinstance(val,list):
        # attachments (list of dicts) or multiselect (list of str)
        if val and isinstance(val[0],dict):
            return '; '.join('%s (%s)'%(a.get('filename'),a.get('url')) for a in val)
        return ', '.join(str(x) for x in val)
    return val

def style_header(ws,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(1,c); cell.font=HEAD; cell.fill=FILL
        cell.alignment=Alignment(vertical='center',horizontal='left',wrap_text=True)
    ws.freeze_panes='A2'
    ws.row_dimensions[1].height=28

def write_records(ws,cols,recs):
    ws.append(cols)
    for i,r in enumerate(recs):
        row=[render(c,r['fields'].get(c)) for c in cols]
        ws.append(row)
    style_header(ws,len(cols))
    # widths
    for ci,c in enumerate(cols,1):
        lens=[len(str(c))]+[min(len(str(render(c,r['fields'].get(c)))),40) for r in recs[:200]]
        ws.column_dimensions[get_column_letter(ci)].width=min(max(max(lens)+2,10),42)
    for ri in range(2,len(recs)+2):
        for ci in range(1,len(cols)+1):
            cell=ws.cell(ri,ci); cell.font=BODY; cell.alignment=TOP; cell.border=BORD
        if ri%2==0:
            for ci in range(1,len(cols)+1): ws.cell(ri,ci).fill=ALT

wb=Workbook()

# Overview
ov=wb.active; ov.title='Overview'
ov['A1']='JobTracker — Airtable Export'; ov['A1'].font=Font(name='Arial',bold=True,size=16)
info=[
 ('Source base','JobTracker2.0.xlsx (Airtable)'),
 ('Base ID','appBhb42sEosIY7ie'),
 ('Table','JobTracker (tbl7yUDz9Z76PNd69)'),
 ('Records',len(records)),
 ('Fields',len(fields)),
 ('Views','Full List View (all fields), Creative Status (filtered)'),
 ('Attachments','1 file — see attachments/ folder'),
 ('Primary field',data['table']['primaryField']),
]
r=3
for k,v in info:
    ov.cell(r,1,k).font=Font(name='Arial',bold=True,size=10)
    ov.cell(r,2,v).font=BODY; r+=1
ov.column_dimensions['A'].width=20; ov.column_dimensions['B'].width=70

# Fields sheet
fs=wb.create_sheet('Fields')
fs.append(['Field Name','Type','Primary','Options (for select fields)'])
for f in fields:
    fs.append([f['name'],f['type'],'Yes' if f.get('isPrimary') else '',
               ', '.join(f['choices']) if f.get('choices') else ''])
style_header(fs,4)
for w,ci in zip([26,18,10,80],[1,2,3,4]): fs.column_dimensions[get_column_letter(ci)].width=w
for ri in range(2,len(fields)+2):
    for ci in range(1,5):
        fs.cell(ri,ci).font=BODY; fs.cell(ri,ci).alignment=WRAP; fs.cell(ri,ci).border=BORD
    if ri%2==0:
        for ci in range(1,5): fs.cell(ri,ci).fill=ALT

# Full List View
flv=wb.create_sheet('Full List View')
write_records(flv,fnames,records)

# Creative Status view (filtered)
cs=[v for v in views if v['name']=='Creative Status'][0]
cs_cols=cs['visibleColumns']
st=set(cs['filter']['conditions'][0]['values']); pt=set(cs['filter']['conditions'][1]['values'])
cs_recs=[r for r in records if r['fields'].get('Project Status') in st and r['fields'].get('Project Type') in pt]
csw=wb.create_sheet('Creative Status')
write_records(csw,cs_cols,cs_recs)

# Views sheet
vs=wb.create_sheet('Views')
vs.append(['View','Type','Visible Columns','Filter','Sort','Group'])
for v in views:
    filt=''
    if v.get('filter'):
        filt=(' %s '%v['filter']['conjunction'].upper()).join(
            '%s %s [%s]'%(c['field'],c['operator'],', '.join(c['values'])) for c in v['filter']['conditions'])
    vs.append([v['name'],v['type'],', '.join(v['visibleColumns']),filt or 'None',v.get('sort') or 'None',v.get('group') or 'None'])
style_header(vs,6)
for w,ci in zip([18,10,60,55,12,12],range(1,7)): vs.column_dimensions[get_column_letter(ci)].width=w
for ri in range(2,len(views)+2):
    for ci in range(1,7):
        vs.cell(ri,ci).font=BODY; vs.cell(ri,ci).alignment=WRAP; vs.cell(ri,ci).border=BORD

wb.save('JobTracker_export.xlsx')
print('Saved. Full List rows:',len(records),'Creative Status rows:',len(cs_recs))
